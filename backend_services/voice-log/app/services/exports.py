"""PDF and Excel export generation from live ADL records."""

from __future__ import annotations

import io
import re
from datetime import date, datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

BRAND_BLUE = colors.HexColor("#1e3a8a")
BRAND_ACCENT = colors.HexColor("#14b8a6")
SLATE_700 = colors.HexColor("#334155")
SLATE_500 = colors.HexColor("#64748b")
ROW_ALT = colors.HexColor("#f1f5f9")
HEADER_HEIGHT = 78

EXCEL_BRAND_BLUE = "1E3A8A"
EXCEL_BRAND_ACCENT = "14B8A6"
EXCEL_SLATE_700 = "334155"
EXCEL_SLATE_500 = "64748B"
EXCEL_ROW_ALT = "F1F5F9"
EXCEL_WHITE = "FFFFFF"
EXCEL_BORDER_COLOR = "CBD5E1"

EXCEL_COLUMNS = [
    ("Date", 12),
    ("Time", 9),
    ("Category", 14),
    ("Transcript", 48),
    ("Alert", 8),
    ("Medication", 16),
    ("Dosage", 12),
    ("Intake level", 14),
]


def _format_time(value: Any) -> str:
    if isinstance(value, str) and "T" in value:
        return value.split("T")[1][:5]
    return str(value)


def _format_record_date(value: Any) -> str:
    if isinstance(value, str) and len(value) >= 10:
        return value[:10]
    return ""


def _clean_summary_line(line: str) -> str:
    text = line.strip()
    if not text:
        return ""
    text = re.sub(r"^═+\s*", "", text)
    text = re.sub(r"\s*═+$", "", text)
    text = text.replace("──", "—")
    return text.strip()


class _ScribePdfDoc(BaseDocTemplate):
    def __init__(self, buffer: io.BytesIO, header_meta: dict[str, str], **kwargs: Any) -> None:
        self.header_meta = header_meta
        super().__init__(buffer, **kwargs)
        frame = Frame(
            18 * mm,
            16 * mm,
            A4[0] - 36 * mm,
            A4[1] - HEADER_HEIGHT - 22 * mm,
            id="main",
        )
        template = PageTemplate(id="main", frames=[frame], onPage=self._draw_header)
        self.addPageTemplates([template])

    def _draw_header(self, canvas: Any, doc: Any) -> None:
        meta = self.header_meta
        width, height = A4
        canvas.saveState()
        canvas.setFillColor(BRAND_BLUE)
        canvas.rect(0, height - HEADER_HEIGHT, width, HEADER_HEIGHT, fill=1, stroke=0)
        canvas.setFillColor(BRAND_ACCENT)
        canvas.rect(0, height - HEADER_HEIGHT - 3, width, 3, fill=1, stroke=0)

        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica-Bold", 17)
        canvas.drawString(18 * mm, height - 28, "SCRIBE ADL Report")

        canvas.setFont("Helvetica", 10)
        canvas.drawString(
            18 * mm,
            height - 44,
            f"Patient {meta['patient_code']}  ·  {meta['period']}",
        )

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 18 * mm, height - 28, meta["generated_date"])
        canvas.drawRightString(
            width - 18 * mm,
            height - 44,
            f"{meta['record_count']} observation(s)",
        )
        canvas.restoreState()


def _excel_thin_border() -> Border:
    side = Side(style="thin", color=EXCEL_BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _style_excel_metadata_row(ws: Any, row_idx: int, *, label_col: int = 1, value_col: int = 2) -> None:
    label_cell = ws.cell(row=row_idx, column=label_col)
    value_cell = ws.cell(row=row_idx, column=value_col)
    label_cell.font = Font(bold=True, color=EXCEL_SLATE_500, size=10)
    label_cell.alignment = Alignment(horizontal="left", vertical="center")
    value_cell.font = Font(color=EXCEL_SLATE_700, size=10)
    value_cell.alignment = Alignment(horizontal="left", vertical="center")


def build_excel_report(
    *,
    patient_code: str,
    start_date: date,
    end_date: date,
    records: list[dict[str, Any]],
    summary_text: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "ADL Report"
    last_col = len(EXCEL_COLUMNS)
    last_col_letter = get_column_letter(last_col)

    title_fill = PatternFill("solid", fgColor=EXCEL_BRAND_BLUE)
    title_font = Font(bold=True, size=16, color=EXCEL_WHITE)
    accent_fill = PatternFill("solid", fgColor=EXCEL_BRAND_ACCENT)
    header_fill = PatternFill("solid", fgColor=EXCEL_BRAND_BLUE)
    header_font = Font(bold=True, size=10, color=EXCEL_WHITE)
    summary_title_fill = PatternFill("solid", fgColor=EXCEL_ROW_ALT)
    summary_title_font = Font(bold=True, size=11, color=EXCEL_BRAND_BLUE)
    border = _excel_thin_border()

    for idx, (_, width) in enumerate(EXCEL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.append(["SCRIBE ADL Report"])
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.append([""])
    ws.merge_cells(f"A2:{last_col_letter}2")
    accent_cell = ws["A2"]
    accent_cell.fill = accent_fill
    ws.row_dimensions[2].height = 4

    metadata_rows = [
        ("Patient", patient_code),
        ("Start date", start_date.isoformat()),
        ("End date", end_date.isoformat()),
        ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Total observations", str(len(records))),
    ]
    for label, value in metadata_rows:
        ws.append([label, value])
        _style_excel_metadata_row(ws, ws.max_row)

    ws.append([])

    headers = [label for label, _ in EXCEL_COLUMNS]
    header_row_idx = ws.max_row + 1
    ws.append(headers)
    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[header_row_idx].height = 22
    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

    data_start_row = header_row_idx + 1
    for row_idx, row in enumerate(records):
        recorded = row.get("recorded_at", "")
        values = [
            _format_record_date(recorded),
            _format_time(recorded),
            (row.get("category") or "").replace("_", " ").title(),
            row.get("cleaned_transcript") or row.get("raw_transcript") or "",
            "Yes" if row.get("alert_required") else "No",
            row.get("medication_name") or "",
            row.get("dosage") or "",
            row.get("intake_level") or "",
        ]
        ws.append(values)
        excel_row = data_start_row + row_idx
        row_fill = PatternFill(
            "solid",
            fgColor=EXCEL_ROW_ALT if row_idx % 2 else EXCEL_WHITE,
        )
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.fill = row_fill
            cell.border = border
            cell.font = Font(color=EXCEL_SLATE_700, size=10)
            if col_idx == 4:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    ws.append([])
    summary_title_row = ws.max_row + 1
    ws.append(["Period Summary"])
    ws.merge_cells(f"A{summary_title_row}:{last_col_letter}{summary_title_row}")
    summary_title_cell = ws[f"A{summary_title_row}"]
    summary_title_cell.font = summary_title_font
    summary_title_cell.fill = summary_title_fill
    summary_title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[summary_title_row].height = 22

    summary_body_row = summary_title_row + 1
    ws.append([summary_text])
    ws.merge_cells(f"A{summary_body_row}:{last_col_letter}{summary_body_row}")
    summary_body_cell = ws[f"A{summary_body_row}"]
    summary_body_cell.font = Font(color=EXCEL_SLATE_700, size=10)
    summary_body_cell.alignment = Alignment(vertical="top", wrap_text=True)
    summary_body_cell.border = border
    estimated_lines = max(4, summary_text.count("\n") + 2)
    ws.row_dimensions[summary_body_row].height = min(estimated_lines * 15, 240)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_pdf_report(
    *,
    patient_code: str,
    start_date: date,
    end_date: date,
    records: list[dict[str, Any]],
    summary_text: str,
    generated_at: datetime | None = None,
) -> bytes:
    generated = generated_at or datetime.now(timezone.utc)
    generated_label = generated.strftime("%d %b %Y · %H:%M UTC")

    buffer = io.BytesIO()
    header_meta = {
        "patient_code": patient_code,
        "period": f"{start_date.isoformat()} to {end_date.isoformat()}",
        "generated_date": f"Generated {generated_label}",
        "record_count": str(len(records)),
    }
    doc = _ScribePdfDoc(
        buffer,
        header_meta,
        pagesize=A4,
        title=f"SCRIBE ADL Report — {patient_code}",
    )

    styles = getSampleStyleSheet()
    body = ParagraphStyle(
        "Body",
        parent=styles["Normal"],
        fontSize=9,
        leading=13,
        textColor=SLATE_700,
        alignment=TA_LEFT,
    )
    section_title = ParagraphStyle(
        "SectionTitle",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=BRAND_BLUE,
        spaceBefore=8,
        spaceAfter=8,
    )
    meta_style = ParagraphStyle(
        "Meta",
        parent=styles["Normal"],
        fontSize=9,
        textColor=SLATE_500,
        spaceAfter=10,
    )
    cell_style = ParagraphStyle(
        "Cell",
        parent=body,
        fontSize=8,
        leading=11,
    )
    header_cell_style = ParagraphStyle(
        "HeaderCell",
        parent=cell_style,
        textColor=colors.white,
    )

    story: list[Any] = []
    story.append(Paragraph("Observation log", section_title))
    story.append(
        Paragraph(
            f"Reporting period <b>{start_date.isoformat()}</b> through "
            f"<b>{end_date.isoformat()}</b>.",
            meta_style,
        )
    )

    if records:
        table_data: list[list[Any]] = [
            [
                Paragraph("<b>Date</b>", header_cell_style),
                Paragraph("<b>Time</b>", header_cell_style),
                Paragraph("<b>Category</b>", header_cell_style),
                Paragraph("<b>Observation</b>", header_cell_style),
                Paragraph("<b>Alert</b>", header_cell_style),
            ]
        ]
        for row in records:
            transcript = (row.get("cleaned_transcript") or row.get("raw_transcript") or "")[:200]
            table_data.append([
                _format_record_date(row.get("recorded_at")),
                _format_time(row.get("recorded_at")),
                (row.get("category") or "").replace("_", " ").title(),
                Paragraph(transcript or "—", cell_style),
                "Yes" if row.get("alert_required") else "No",
            ])

        table = Table(
            table_data,
            colWidths=[58, 42, 72, 255, 38],
            repeatRows=1,
        )
        table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), BRAND_BLUE),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("TEXTCOLOR", (0, 1), (-1, -1), SLATE_700),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ROW_ALT]),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ])
        )
        story.append(table)
    else:
        story.append(Paragraph("No observations were recorded in this period.", body))

    story.append(Spacer(1, 16))
    story.append(Paragraph("Period summary", section_title))

    for raw_line in summary_text.split("\n"):
        line = _clean_summary_line(raw_line)
        if not line or line.lower().startswith("end of"):
            continue
        if line.startswith("Period Summary") or line.startswith("═══"):
            continue
        story.append(Paragraph(line, body))

    doc.build(story)
    return buffer.getvalue()
